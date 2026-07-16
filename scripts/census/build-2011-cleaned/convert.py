import openpyxl, csv, re, os, glob, json, gzip, io
ROOT="/home/user/civgraph/data/census/2011"
CODE=re.compile(r'^\d{1,4}$')
GEOGTAG={"SMALL AREAS":"SA2011","SUPER OUTPUT AREAS":"SOA2011","ELECTORAL WARDS":"WARD2014",
         "ASSEMBLY AREAS":"AA2014","HIGHER_GEOGRAPHIES":"HIGHER","LOCAL GOVERNMENT DISTRICTS":"LGD2014"}

# ---------- geography label lookups (CODE->NAME) per dataset dir ----------
def load_geo(dsdir):
    g={}
    for f in glob.glob(f"{dsdir}/All_Geographies_Code_Files/*.csv"):
        with open(f, encoding='latin-1') as fh:
            rd=csv.reader(fh); hdr=next(rd,None)
            if not hdr or [h.strip().upper() for h in hdr[:2]]!=["CODE","NAME"]:
                continue   # skip hierarchy/correspondence files; keep only CODE,NAME name tables
            for row in rd:
                if len(row)>=2 and row[0].strip(): g[row[0].strip()]=row[1].strip()
    return g

# ---------- outline parsing ----------
def load_outlines(xlsx):
    wb=openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    out={}
    for sn in wb.sheetnames:
        ws=wb[sn]
        rows=[[(str(c).rstrip() if c is not None else '') for c in r] for r in ws.iter_rows(values_only=True)]
        if not rows or not rows[0] or not rows[0][0]: continue
        m=re.match(r'(?:Table\s+)?([A-Z]{2}\d+[A-Z]{2}):\s*(.*)', rows[0][0])
        if not m: continue
        tid=m.group(1); title=re.sub(r'\s*Back to Table list.*','',m.group(2)).strip().rstrip('|').strip()
        out[tid]=(title, rows)
    return out

def is_flat(rows):
    return any(r and r[0].strip()=='Unit of Measure' for r in rows)

def parse_flat(tid,title,rows):
    # QS/KS: category header row = last header row before 'Statistical Unit'; unit row; code row (starts 'Area')
    su=um=area=None
    for i,r in enumerate(rows):
        if r and r[0].strip()=='Statistical Unit': su=i
        if r and r[0].strip()=='Unit of Measure': um=i
        if r and r[0].strip()=='Area': area=i
    cathdr=su-1
    # forward-fill category header
    cat=list(rows[cathdr]); last=''
    for j in range(1,len(cat)):
        if cat[j].strip(): last=cat[j].strip()
        cat[j]=last
    codemap={}
    for j in range(1,len(rows[area])):
        code=rows[area][j].strip()
        if not CODE.match(code): continue
        unit=rows[um][j].strip() if um is not None and j<len(rows[um]) else 'Count'
        if unit.lower()!='count': continue           # drop Percentage / derived
        codemap[tid+code.zfill(4)]=[cat[j] if j<len(cat) else '']
    dim=re.split(r'\bBY\b',title)[0].strip()
    return [dim], codemap

def parse_crosstab(tid,title,rows):
    dims=[d.strip() for d in re.split(r'\bBY\b',title)]
    first=cstart=None
    for i,r in enumerate(rows):
        cs=[j for j,v in enumerate(r) if re.match(r'^\d{4}$',v)]
        if cs: first=i; cstart=min(cs); break
    hdr_rows=[i for i in range(first) if any(rows[i][j].strip() for j in range(cstart,len(rows[i])))]
    def ff(r):
        o=list(r); last=''
        for j in range(cstart,len(o)):
            if o[j].strip(): last=o[j].strip()
            o[j]=last
        return o
    hdr=[ff(rows[i]) for i in hdr_rows]
    codemap={}
    for i in range(first,len(rows)):
        r=rows[i]
        rl=[r[j].strip() for j in range(cstart)]
        if not any(rl): continue
        for j in range(cstart,len(r)):
            if re.match(r'^\d{4}$',r[j]):
                cl=[hdr[k][j] for k in range(len(hdr_rows))]
                codemap[tid+r[j]]=rl+cl   # row dims then col dims (axis order)
    return dims, codemap

def parse(tid,title,rows):
    return parse_flat(tid,title,rows) if is_flat(rows) else parse_crosstab(tid,title,rows)
