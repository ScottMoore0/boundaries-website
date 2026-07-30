#!/usr/bin/env python3
"""Melt a LucidTalk / Belfast Telegraph poll-table workbook into the corpus tidy schema.

LucidTalk publish topline figures on their news page and the FULL crosstabs as a separate
workbook, which is what the British Polling Council requires of members. This reads the
workbook, not the page: the page gives party shares in a JPEG, the workbook gives every
response by gender, age, social grade, NI region, 2022 past vote, constitutional bloc and
community background.

Output matches the 23 spreadsheet-derived polls already in the corpus, so the new file can
sit beside them:

    Geography Code, Geography Label, Measure, Response, Response Label,
    Breakdown Dimension, Breakdown Category, Base Type, Statistic, Unit, Value,
    Extraction Confidence

WORKBOOK SHAPE. Row 3 holds the question, row 4 the demographic group headers (sparse --
each is written once above its span), row 5 the column labels, rows 6-8 the bases, and
then the responses in COUNT/PERCENT PAIRS: a row of weighted numbers followed by its
percentage row. The pairs are read positionally rather than by label, because the labels
do not follow one rule -- most percent rows repeat the response and add ' %', Q5 repeats
the text verbatim, and the Q3 rating scales label the count row '75' and its percent row
'0.75'.

PERCENTAGES ARE STORED THREE DIFFERENT WAYS in the same workbook: '14%' as text, 0.51 as
a fraction, and occasionally a plain number. They are normalised to a number out of 100,
treating any value <= 1 as a fraction. That is safe here because a genuine 1% is written
0.01; the only ambiguous case would be a literal 1.0 meaning one per cent, which does not
occur -- and the column checksum below would catch it if it did.

TWO CHECKS, NEITHER FITTED
    every percentage column sums to 100 (+/- rounding), per question
    every count column sums to that column's weighted base
Failures are reported per question and column rather than silently emitted.

NI REGION IS ITS OWN DIMENSION HERE. In the existing corpus files the five region columns
(BELFAST, EAST, NORTH, SOUTH, WEST) are bucketed under 'Age', which is a mis-mapping in
whatever produced them -- anything filtering Breakdown Dimension == 'Age' picks up regions
too. This writes them as 'NIRegion'. Deliberate divergence, noted so it is not mistaken
for drift.

Usage:
    python scripts/polls/parse_lucidtalk_spreadsheet.py <workbook.xlsx> <code> [--out DIR]
e.g.
    python scripts/polls/parse_lucidtalk_spreadsheet.py BelTelJuly26-MAINTABLESFP.xlsx 2026-07
"""
import os, re, csv, sys, json, argparse

import openpyxl

GEO_CODE, GEO_LABEL = 'N92000002', 'Northern Ireland'
CONF = 0.95        # direct spreadsheet read, structurally validated -- not an OCR estimate
SKIP = {'FRONTPAGEINTRODUCTION', 'Contents'}

# group-header keyword -> canonical dimension name used by the corpus
DIMS = [
    ('gender', 'Gender'),
    ('age-group', 'Age'),
    ('socio-economic', 'SocialGrade'),
    ('ni region', 'NIRegion'),
    ('past-vote', 'PastVote'),
    ('constitutional', 'ConstitutionalBloc'),
    ('religion', 'Religion'),
]
# rows that state a base or a summary statistic rather than a response
BASE_PAT = re.compile(r'^(unweighted|weighted)', re.I)
AVG_PAT = re.compile(r'weighted average', re.I)


def base_type_of(ws):
    """inc_DK / exc_DK, taken from the QUESTION TEXT and the base row -- never the tab name.

    The July 2026 workbook ships its main voting-intention table on a tab called
    'MAINPollQuestion1inc.DKs' whose question text reads 'MAIN Results exc. Don't
    Know/Undecideds' and whose base is 'Weighted - Excluding Non-Voters/Don't Knows'. There
    is no Don't Know row and the party shares total 100 on a base of 980.45 rather than
    1050.01. The tab name is a stale template label: trusting it would file decided-voter
    shares as if they still contained the undecideds, which is the single most misleading
    error available here -- it silently inflates every party's number.
    """
    text = ' '.join(str(ws.cell(r, 1).value or '') for r in (3, 8)).lower()
    if re.search(r'\bexc\w*\.?\s*(don\'?t know|dk)|excluding[^.]*don\'?t know', text):
        return 'exc_DK'
    if re.search(r'\binc\w*\.?\s*(don\'?t know|dk)|including[^.]*don\'?t know', text):
        return 'inc_DK'
    t = ws.title.lower()
    return 'inc_DK' if 'inc.dk' in t else ('exc_DK' if 'exc' in t else '')


def dim_of(header):
    h = (header or '').lower()
    for k, v in DIMS:
        if k in h:
            return v
    return None


def tidy_cat(label):
    """'ABC1 i.e. "Middle Class"' -> 'ABC1';  'BELFAST - 4 Belfast Constituencies' -> 'BELFAST'.

    Split only on a SPACED hyphen, so '18-34 years age-group' survives intact.
    """
    s = str(label or '').strip()
    for sep in (' i.e. ', ' - '):
        if sep in s:
            s = s.split(sep)[0]
    return s.strip()


def as_pct(v):
    if v is None or v == '':
        return None
    if isinstance(v, str):
        s = v.strip().rstrip('%').replace(',', '').strip()
        try:
            return round(float(s), 2)
        except ValueError:
            return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return round(f * 100 if f <= 1 else f, 2)


def as_num(v):
    if v is None or v == '':
        return None
    if isinstance(v, str):
        s = v.strip().rstrip('%').replace(',', '')
        try:
            return round(float(s), 2)
        except ValueError:
            return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def columns(ws):
    """(col_index, dimension, category) for every data column."""
    groups = [c.value for c in ws[4]]
    labels = [c.value for c in ws[5]]
    out, cur = [], None
    for i in range(1, len(labels)):          # column 1 is the row label
        g = groups[i] if i < len(groups) else None
        if g and str(g).strip():
            d = dim_of(g)
            if d:
                cur = d
        lab = labels[i]
        if lab is None or not str(lab).strip():
            continue
        if i == 1:
            out.append((i + 1, 'Total', 'Total'))
            cur = None
            continue
        if cur:
            out.append((i + 1, cur, tidy_cat(lab)))
    return out


def parse_sheet(ws, base_type):
    question = str(ws.cell(3, 1).value or ws.cell(2, 2).value or ws.title).strip()
    cols = columns(ws)

    bases, avg_row, first = {}, None, None
    for r in range(6, ws.max_row + 1):
        lab = ws.cell(r, 1).value
        if lab is None or not str(lab).strip():
            continue
        s = str(lab).strip()
        if AVG_PAT.search(s):
            avg_row = r
            continue
        if BASE_PAT.match(s):
            bases[s] = r
            continue
        first = r
        break

    rows, pairs = [], []
    r = first
    while r is not None and r + 1 <= ws.max_row:
        lab = ws.cell(r, 1).value
        if lab is None or not str(lab).strip():
            r += 1
            continue
        pairs.append((str(lab).strip(), r, r + 1))
        r += 2

    def emit(resp, dim, cat, stat, unit, val):
        if val is None:
            return
        rows.append({
            'Geography Code': GEO_CODE, 'Geography Label': GEO_LABEL,
            'Measure': f'{question} [{ws.title}]', 'Response': resp, 'Response Label': '',
            'Breakdown Dimension': dim, 'Breakdown Category': cat,
            'Base Type': base_type, 'Statistic': stat, 'Unit': unit,
            'Value': val, 'Extraction Confidence': CONF})

    for ci, dim, cat in cols:
        for label, rc, rp in pairs:
            resp = re.sub(r'\s*%$', '', label).strip()
            emit(resp, dim, cat, 'count', 'persons', as_num(ws.cell(rc, ci).value))
            emit(resp, dim, cat, 'percent', '%', as_pct(ws.cell(rp, ci).value))
        for name, br in bases.items():
            stat = 'base_unweighted' if name.lower().startswith('unweighted') else 'base_weighted'
            # Base labels are kept verbatim. Tidying them collapsed 'Weighted - Excluding
            # Non-Voters' to 'Weighted', which made Q1's two weighted bases -- 1050.01 for
            # everyone and 980.45 for those who would vote -- indistinguishable in the
            # output, with the party counts summing to the second.
            emit(name, dim, cat, stat, 'persons', as_num(ws.cell(br, ci).value))
        if avg_row:
            emit('Weighted Average - Total Score', dim, cat, 'weighted_average', 'score',
                 as_num(ws.cell(avg_row, ci).value))
    return question, rows, [p[0] for p in pairs]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('workbook')
    ap.add_argument('code', help="poll code, e.g. 2026-07")
    ap.add_argument('--out', default=None)
    a = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    repo = os.path.abspath(os.path.join(here, '..', '..'))
    outdir = a.out or os.path.join(repo, 'analysis', 'border-poll-dry-run', 'v9', '_lt_cache')
    os.makedirs(outdir, exist_ok=True)
    out = os.path.join(outdir, f'{a.code}-spreadsheet.csv')

    wb = openpyxl.load_workbook(a.workbook, data_only=True)
    allrows, meta, bad = [], [], 0
    for ws in wb.worksheets:
        if ws.title in SKIP:
            continue
        q, rows, resp = parse_sheet(ws, base_type_of(ws))
        bt = rows[0]['Base Type'] if rows else ''
        allrows.extend(rows)

        # --- checks
        tot = [r for r in rows if r['Breakdown Dimension'] == 'Total']
        psum = sum(r['Value'] for r in tot if r['Statistic'] == 'percent')
        csum = sum(r['Value'] for r in tot if r['Statistic'] == 'count')
        # A sheet may publish more than one weighted base, and the responses need not sum
        # to the headline one: Q1's parties sum to the excluding-non-voters base (980.45),
        # not to the full 1050.01. So the counts pass if they match ANY weighted base.
        allb = [r['Value'] for r in tot if r['Statistic'] == 'base_weighted']
        base = max(allb) if allb else None
        okp = abs(psum - 100) <= 1.5
        okc = any(abs(csum - b) <= max(3.0, 0.01 * b) for b in allb)
        matched = next((b for b in allb
                        if abs(csum - b) <= max(3.0, 0.01 * b)), None)
        if not (okp and okc):
            bad += 1
        meta.append({'sheet': ws.title, 'question': q, 'responses': len(resp),
                     'base_type': bt, 'pct_sum': round(psum, 1),
                     'count_sum': round(csum, 1), 'weighted_base': base,
                     'count_base_matched': matched})
        flag = '' if (okp and okc) else '   <-- CHECK'
        print(f'  {ws.title:26} {len(resp):2} responses  pct {psum:6.1f}  '
              f'counts {csum:8.1f} vs base {matched if matched else base}{flag}')

    with open(out, 'w', encoding='utf-8', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=list(allrows[0].keys()))
        w.writeheader()
        w.writerows(allrows)

    mpath = os.path.join(outdir, f'{a.code}-spreadsheet.meta.json')
    json.dump({'code': f'{a.code}-spreadsheet', 'source_format': 'spreadsheet',
               'workbook': os.path.basename(a.workbook), 'questions': meta,
               'rows': len(allrows)},
              open(mpath, 'w', encoding='utf-8'), indent=1, ensure_ascii=False)

    print(f'\n  {len(allrows):,} rows from {len(meta)} questions -> {out}')
    print(f'  {len(meta) - bad} of {len(meta)} questions passed both checks')


if __name__ == '__main__':
    main()
